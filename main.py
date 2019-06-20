import os
import openpyxl.styles as sty
import openpyxl
import xlrd
import error
import utils
import global_var as gl
import re
import shutil


def execute(csv_dir, index_file, trans_file, version):
    if csv_dir == "" or index_file == "" or trans_file == "":
        error.Error.set_code(0, "csv_dir = " + csv_dir + ", index_file = " + index_file + ", trans_file = " + trans_file)
        return
    trans_dir = trans_file.replace('.xlsx', '')
    # read index file, csv
    try:
        arr_index = utils.read_csv(index_file)
    except Exception as e:
        error.Error.set_code(4, str(e))
        return
    if not arr_index:
        error.Error.set_code(0, "index_file is None")
        return
    # back up old file
    try:
        shutil.copy(trans_file, trans_dir+'_old.xlsx')
    except Exception as e:
        error.Error.set_code(6, str(e))
        return
    # read trans file, xls
    try:
        old_xls = xlrd.open_workbook(trans_file)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    old_sheet = old_xls.sheet_by_index(0)
    new_xls = openpyxl.Workbook()
    new_sheet = new_xls.active

    # construct dict shared id, dict value
    # dict_id = {id: [sid, lang_key, designer, system]}
    dict_id = {}
    # dict_row = {id: row}
    dict_row = {}
    # dict_lang_key = {lang_key: id}
    dict_lang_key = {}
    # tar_dict_id = {tar_id: [sid, lang_key, color, designer, system]}
    tar_dict_id = {}
    # tar_dict_line_index = {tar_id: virtual_line}
    tar_dict_line_index = {}
    # tar_key_line = [virtual_line1 = [tar_id1, tar_id3], v_line2 = [tar_id2] ... ] for sorting of records
    tar_key_line = []
    number_id = 0
    for i in range(old_sheet.nrows):
        if i < gl.trans_content_row:
            # read number id
            if i == 1:
                number_id = int(old_sheet.cell_value(i, gl.col_nid))
            for title_col in range(old_sheet.ncols):
                new_sheet.cell(row=i+1, column=title_col+1, value=old_sheet.cell_value(i, title_col))
            continue
        ref_id = old_sheet.cell_value(i, gl.col_id)
        ref_sid = old_sheet.cell_value(i, gl.col_sid)
        ref_lang_key = old_sheet.cell_value(i, gl.col_langkey)
        dict_id[ref_id] = []
        dict_id[ref_id].append(ref_sid)
        dict_id[ref_id].append(ref_lang_key)
        dict_row[ref_id] = i
        if ref_lang_key != "":
            dict_lang_key[ref_lang_key] = ref_id
    # read csv data through index file
    for i in range(len(arr_index)):
        if i < gl.index_content_row:
            orig_name = ""
            continue
        if i > gl.index_content_row:
            # optimization for reducing open operation of the same csv file
            orig_name = arr_index[i-1][0]
        csv_name = arr_index[i][gl.index_csv_name]
        attr = arr_index[i][gl.index_csv_col]
        designer = arr_index[i][gl.index_designer]
        system = arr_index[i][gl.index_sys]
        print("Processing: " + csv_name + ".csv:  " + attr)
        # open csv file
        if csv_name != orig_name:
            csv_path = csv_dir + "/" + csv_name + ".csv"
            try:
                arr_csv = utils.read_csv(csv_path)
            except Exception as e:
                error.Error.set_code(5, csv_name+" "+str(e))
                return
            if not arr_csv:
                error.Error.set_code(1, csv_name)
                return
        # find attr col of attr
        for col in range(len(arr_csv[gl.field_row])):
            if arr_csv[gl.field_row][col] == attr:
                attr_col = col
                break
        # find key col list
        key_col = []
        for col in range(len(arr_csv[gl.key_row])):
            if "key" in arr_csv[gl.key_row][col].lower().split(';'):
                key_col.append(col)
        constant_text_flag = False
        if csv_name.lower() == gl.constant_text_file:
            constant_text_flag = True
        # traverse all records
        is_array_string = False
        for row in range(len(arr_csv)):
            if row == gl.type_row:
                if arr_csv[row][attr_col].lower().strip() == gl.array_string_pattern:
                    is_array_string = True
                continue
            if row < gl.content_row:
                continue
            # if csv is ConstantText, read designer, system from the file
            if constant_text_flag:
                designer = arr_csv[row][gl.constant_text_designer]
                system = arr_csv[row][gl.constant_text_sys]

            val_list = utils.split_array_string(arr_csv[row][attr_col], is_array_string)

            for val_index in range(len(val_list)):
                # construct target id, get tar_val
                tar_id = csv_name + "#" + attr
                for k_col in key_col:
                    tar_id = tar_id + "#" + arr_csv[row][k_col]
                if is_array_string:
                    tar_id = tar_id + str(val_index)
                tar_val = val_list[val_index]
                if tar_val == "":
                    continue
                # construct tar_dict_val and tar_dict_sid
                is_repeat = False
                for k, v in tar_dict_id.items():
                    if v[1] == tar_val and v[0] == "":
                        # init tar_dict_id[tar_id] = [sid, lang_key, ref_id, color, designer, system]
                        tar_dict_id[tar_id] = ["", "", "", "", "", ""]
                        tar_dict_id[tar_id][0] = k
                        tar_dict_id[tar_id][1] = ""
                        line_index = tar_dict_line_index[k]
                        tar_key_line[line_index].append(tar_id)
                        is_repeat = True
                        break
                if not is_repeat:
                    # init tar_dict_id[tar_id] = [sid, lang_key, ref_id, color, designer, system]
                    tar_dict_id[tar_id] = ["", "", "", "", "", ""]
                    tar_dict_id[tar_id][0] = ""
                    tar_dict_id[tar_id][1] = tar_val
                    tar_dict_line_index[tar_id] = len(tar_key_line)
                    tar_key_line.append([tar_id])
                    # set parameters
                    tar_dict_id[tar_id][4] = designer
                    tar_dict_id[tar_id][5] = system

                # compare tar_dict_val/tar_dict_sid with dict_val/dict_sid
                if dict_id.__contains__(tar_id):
                    # 1. modify
                    if not (dict_id[tar_id][1] == tar_dict_id[tar_id][1] and dict_id[tar_id][0] == tar_dict_id[tar_id][0]):
                        if dict_lang_key.__contains__(tar_dict_id[tar_id][1]):
                            tar_dict_id[tar_id][2] = dict_lang_key[tar_dict_id[tar_id][1]]
                        tar_dict_id[tar_id][3] = gl.color_modify
                    # 2. ID and LangKey are identical
                    else:
                        tar_dict_id[tar_id][2] = tar_id
                        mark_green_flag = False
                        if old_sheet.cell_value(dict_row[tar_id], gl.col_langkey+3) != "":
                            mark_green_flag = True
                        if mark_green_flag:
                            tar_dict_id[tar_id][3] = gl.color_fill
                        else:
                            tar_dict_id[tar_id][3] = gl.color_default
                # 3. add new records
                else:
                    if dict_lang_key.__contains__(tar_dict_id[tar_id][1]):
                        tar_dict_id[tar_id][2] = dict_lang_key[tar_dict_id[tar_id][1]]
                    tar_dict_id[tar_id][3] = gl.color_add

    content_row = gl.trans_content_row
    for key_list in tar_key_line:
        for key_id in key_list:
            for col in range(old_sheet.ncols):
                fill_type = sty.PatternFill(fill_type=None)
                if tar_dict_id[key_id][3] != gl.color_default:
                    fill_type = sty.PatternFill(fill_type='solid', fgColor=tar_dict_id[key_id][3])
                if col == gl.col_id:
                    new_sheet.cell(content_row+1, col+1, key_id).fill = fill_type
                elif col == gl.col_sid:
                    new_sheet.cell(content_row+1, col+1, tar_dict_id[key_id][0]).fill = fill_type
                elif col == gl.col_nid:
                    if tar_dict_id[key_id][3] == gl.color_add or tar_dict_id[key_id][3] == gl.color_modify:
                        if tar_dict_id[key_id][0] == "":
                            new_sheet.cell(content_row+1, col+1, number_id+1).fill = fill_type
                            number_id += 1
                        else:
                            new_sheet.cell(content_row+1, col+1, "").fill = fill_type
                    else:
                        new_sheet.cell(content_row+1, col+1,
                                       old_sheet.cell_value(dict_row[key_id], col)).fill = fill_type
                elif col == gl.col_ignore:
                    if tar_dict_id[key_id][0] != "":
                        new_sheet.cell(content_row+1, col+1, "Share ID").fill = fill_type
                    else:
                        tmp_old_ignore = ""
                        if tar_dict_id[key_id][3] == gl.color_modify:
                            tmp_old_ignore = old_sheet.cell_value(dict_row[key_id], col)
                        if tar_dict_id[key_id][2] != "":
                            tmp_old_ignore = old_sheet.cell_value(dict_row[tar_dict_id[key_id][2]], col)
                        # tricky bug-fix
                        if tmp_old_ignore == "Share ID":
                            tmp_old_ignore = ""
                        new_sheet.cell(content_row+1, col+1, tmp_old_ignore).fill = fill_type
                elif col == gl.col_hist:
                    if tar_dict_id[key_id][3] == gl.color_add:
                        if version != "":
                            new_sheet.cell(content_row+1, col+1, version).fill = fill_type
                        else:
                            new_sheet.cell(content_row+1, col+1, "").fill = fill_type
                    elif tar_dict_id[key_id][3] == gl.color_modify:
                        tmp_value = old_sheet.cell_value(dict_row[key_id], col)
                        new_sheet.cell(content_row+1, col+1,
                                       utils.combine_version(tmp_value, version)).fill = fill_type
                    else:
                        new_sheet.cell(content_row+1, col+1,
                                       old_sheet.cell_value(dict_row[key_id], col)).fill = fill_type
                elif col == gl.col_designer:
                    new_sheet.cell(content_row+1, col+1, tar_dict_id[key_id][4]).fill = fill_type
                elif col == gl.col_sys:
                    new_sheet.cell(content_row+1, col+1, tar_dict_id[key_id][5]).fill = fill_type
                elif col == gl.col_feature or col == gl.col_term or col == gl.col_desc or col == gl.col_instruction:
                    if tar_dict_id[key_id][0] != "":
                        new_sheet.cell(content_row+1, col+1, "").fill = fill_type
                    else:
                        tmp_old_elem = ""
                        if tar_dict_id[key_id][3] == gl.color_modify:
                            tmp_old_elem = old_sheet.cell_value(dict_row[key_id], col)
                        if tar_dict_id[key_id][2] != "":
                            tmp_old_elem = old_sheet.cell_value(dict_row[tar_dict_id[key_id][2]], col)
                        new_sheet.cell(content_row+1, col+1, tmp_old_elem).fill = fill_type
                elif col == gl.col_langkey:
                    new_sheet.cell(content_row+1, col+1, tar_dict_id[key_id][1]).fill = fill_type
                elif col == gl.col_langkey+1:
                    if tar_dict_id[key_id][3] == gl.color_add or tar_dict_id[key_id][3] == gl.color_modify:
                        re_str = re.sub(gl.chs_pattern, "", tar_dict_id[key_id][1])
                        new_sheet.cell(content_row+1, col+1, re_str).fill = fill_type
                    else:
                        tmp_lang = old_sheet.cell_value(dict_row[tar_dict_id[key_id][2]], col)
                        new_sheet.cell(content_row+1, col+1, tmp_lang).fill = fill_type
                else: # col > gl.col_langkey+1:
                    if tar_dict_id[key_id][2] == "":
                        new_sheet.cell(content_row+1, col+1, "").fill = fill_type
                    else:
                        tmp_lang = old_sheet.cell_value(dict_row[tar_dict_id[key_id][2]], col)
                        new_sheet.cell(content_row+1, col+1, tmp_lang).fill = fill_type

            content_row += 1

    # write latest number_id
    new_sheet.cell(row=gl.max_num_id_col+1, column=gl.col_nid+1, value=number_id)
    new_sheet.freeze_panes = "A2"
    try:
        os.remove(trans_file)
    except Exception as e:
        error.Error.set_code(2, str(e))
        return

    try:
        new_xls.save(trans_file)
    except Exception as e:
        error.Error.set_code(7, str(e))
        return

    new_xls.close()


def gen_task(trans_file, task_file, version):
    if trans_file == "" or task_file == "":
        error.Error.set_code(0, "trans/task")
        return
    task_file = task_file.replace('$version', version)
    task_dir = task_file.replace('.xlsx', '')
    if os.path.exists(task_file):
        if os.path.exists(task_dir+"_old.xlsx"):
            try:
                os.remove(task_dir+"_old.xlsx")
            except Exception as e:
                error.Error.set_code(2, str(e))
                return
        os.rename(task_file, task_dir+"_old.xlsx")
    # read trans file, xls
    try:
        trans_xls = xlrd.open_workbook(trans_file)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    trans_sheet = trans_xls.sheet_by_index(0)
    task_xls = openpyxl.Workbook()
    task_sheet = task_xls.active

    # fill title
    task_cols = 0
    for i in range(trans_sheet.ncols):
        if i == gl.col_ignore or i == gl.col_designer:
            continue
        task_sheet.cell(1, task_cols+1, trans_sheet.cell_value(0, i))
        task_cols += 1
    task_cols -= 1

    # traverse all records
    row_index = 1
    for i in range(trans_sheet.nrows):
        if i < gl.trans_content_row:
            continue
        if trans_sheet.cell_value(i, gl.col_sid) != "":
            continue
        if trans_sheet.cell_value(i, gl.col_ignore) != "":
            continue
        col_index = 0
        fix_lang = []
        is_trans_finished = True
        for j in range(trans_sheet.ncols):
            fill_type = sty.PatternFill(fill_type=None)
            if j == gl.col_ignore or j == gl.col_designer:
                continue
            if j == gl.col_instruction:
                instruction = trans_sheet.cell_value(i, j)
                result = re.findall(gl.instruction_pattern, instruction)
                if result:
                    fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_modify)
                for m in result:
                    if m[0].lower().strip() == "fix":
                        fix_lang.append(m[1].lower().strip())
            elif j > gl.col_langkey + 1 and j != gl.col_langkey + 3:
                if trans_sheet.cell_value(i, j) == "":
                    fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_fill)
                    is_trans_finished = False
                if fix_lang.__contains__("all") or fix_lang.__contains__(trans_sheet.cell_value(0, j).lower()):
                    fill_type = sty.PatternFill(fill_type='solid', fgColor=gl.color_modify)
                    is_trans_finished = False

            task_sheet.cell(row_index+1, col_index+1,
                            trans_sheet.cell_value(i, j)).fill = fill_type
            col_index += 1

        if not is_trans_finished:
            row_index += 1

    task_sheet.freeze_panes = "A2"
    try:
        task_xls.save(task_file)
    except Exception as e:
        error.Error.set_code(7, str(e))
        return
    task_xls.close()


def update_glossary(trans_file, glossary_file, version):
    if trans_file == "" or glossary_file == "":
        error.Error.set_code(0, "trans/glossary")
        return
    # glossary_file = glossary_file.replace('$version', version)
    glossary_dir = glossary_file.replace('.xlsx', '')
    if os.path.exists(glossary_file):
        if os.path.exists(glossary_dir+"_old.xlsx"):
            try:
                os.remove(glossary_dir+"_old.xlsx")
            except Exception as e:
                error.Error.set_code(2, str(e))
                return
        os.rename(glossary_file, glossary_dir+"_old.xlsx")
    # read trans file, xls
    try:
        trans_xls = xlrd.open_workbook(trans_file)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    trans_sheet = trans_xls.sheet_by_index(0)
    glossary_xls = openpyxl.Workbook()
    glossary_sheet = glossary_xls.active

    # fill title
    glossary_cols = 0
    for i in range(trans_sheet.ncols):
        if i == gl.col_hist or i == gl.col_term or i > gl.col_langkey:
            glossary_sheet.cell(1, glossary_cols+1, trans_sheet.cell_value(0, i))
            glossary_cols += 1
    glossary_cols -= 1

    # save all term and id
    term_id_list = []
    dict_row = {}
    for i in range(trans_sheet.nrows):
        if i < gl.trans_content_row:
            continue
        if trans_sheet.cell_value(i, gl.col_term) == "":
            continue
        trans_id = trans_sheet.cell_value(i, gl.col_id)
        trans_term = trans_sheet.cell_value(i, gl.col_term)
        term_id_list.append(trans_term+'+'+trans_id)
        dict_row[trans_id] = i
    # sort term by alphabetic order
    #term_id_list.sort()
    term_id_list = sorted(term_id_list, key=str.lower)
    # fill sheet
    row_index = 1
    for i in range(len(term_id_list)):
        id_name = term_id_list[i].split('+')[1]
        col_index = 0
        for j in range(trans_sheet.ncols):
            if j == gl.col_hist or j == gl.col_term or j > gl.col_langkey:
                glossary_sheet.cell(row_index+1, col_index+1,
                                    trans_sheet.cell_value(dict_row[id_name], j))
                col_index += 1
        row_index += 1

    glossary_sheet.freeze_panes = "A2"
    try:
        glossary_xls.save(glossary_file)
    except Exception as e:
        error.Error.set_code(7, str(e))
        return
    glossary_xls.close()


def convert(lua_dir, output_file):
    if lua_dir == "" or output_file == "":
        error.Error.set_code(0, "lua_dir/output")
        return

    try:
        c_excel = xlrd.open_workbook(output_file)
    except Exception as e:
        error.Error.set_code(3, str(e))
        return
    c_sheet = c_excel.sheet_by_index(0)

    fp = open(lua_dir+"/TransTable.lua", 'w', encoding="UTF-8")
    fp.write("-- ********************************************************\n")
    fp.write("-- generated by csv utility, please do not change manually\n")
    fp.write("-- ********************************************************\n")
    fp.write("Table = { \n")

    chs_col = 0
    eng_col = 0

    dict_sid = {}
    dict_val = {}
    dict_row = {}

    for i in range(c_sheet.nrows):
        if i < 4:
            continue
        dict_sid[c_sheet.cell_value(i, gl.col_id)] = c_sheet.cell_value(i, gl.col_sid)
        dict_row[c_sheet.cell_value(i, gl.col_id)] = i

    for i in range(c_sheet.nrows):
        if i == 0:
            for j in range(c_sheet.ncols):
                if c_sheet.cell_value(i, j) == "CHS":
                    chs_col = j
                elif c_sheet.cell_value(i, j) == "EN":
                    eng_col = j
            continue
        if i < 4:
            continue
        fp.write("[\""+c_sheet.cell_value(i, gl.col_id) + "\"] = {\n")
        fp.write("    id = \""+c_sheet.cell_value(i, gl.col_id) + "\",\n")
        right_line = utils.get_val_line(dict_sid, dict_row, i)
        fp.write("    CHS = \""+c_sheet.cell_value(right_line, gl.col_langkey+1) + "\",\n")
        fp.write("    ENG = \""+c_sheet.cell_value(right_line, gl.col_langkey+3) + "\",\n")
        fp.write("},\n")

    fp.write("}\n")
    fp.write("return Table\n")
    fp.close()


def export_csv(trans_file, glossary_file):
    if not os.path.exists(glossary_file) or not os.path.exists(trans_file):
        error.Error.set_code(0, "trans/glossary")
        return
    utils.xlsx_to_csv(trans_file)
    utils.xlsx_to_csv(glossary_file)
